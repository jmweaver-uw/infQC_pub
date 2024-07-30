import numpy as np
import os
from custom_data_gen import DataGenerator
from openpyxl import Workbook
import json
from matplotlib import pyplot as plt
import tensorflow as tf
from data_processing.genSplits import create_partitions
keras = tf.keras
from keras.layers import *
from keras.callbacks import EarlyStopping, ModelCheckpoint

# Parameters
trainparams = {'dim': (128, 128, 70),
               'batch_size': 8,
               'n_classes': 2,
               'n_channels': 1,
               'shuffle': True}

testparams = {'dim': (128, 128, 70),
              'batch_size': 8,
              'n_classes': 2,
              'n_channels': 1,
              'shuffle': True}

labels_json_path = ' '
folds_json_path = ' '
volumes_json_path = ' '
model_save_path = ' '

# open worksheet
wb = Workbook()
ws_loss = wb.active
ws_loss.title = "Loss"
ws_acc = wb.create_sheet(title="Acc")
ws_loss.cell(1, 1).value = 'epoch#'
ws_acc.cell(1, 1).value = 'epoch#'

# load labels
f = open(labels_json_path)
labels = json.load(f)
f.close()

# load fold dictionary
f = open(folds_json_path)
foldDict = json.load(f)
f.close()

# load subVol dictionary
f = open(volumes_json_path)
subVols = json.load(f)
f.close()

# Loop through folds, switch "for" statement if only running one fold, i.e., for fold in range(4, 5)
for fold in range(1, len(foldDict)+1):
    print("Training fold # " + str(fold) + "/" + str(len(foldDict)))

    # CREATE PARTITION
    partition = create_partitions(labels, foldDict, subVols, fold)
    # Generators
    training_generator = DataGenerator(partition['train'], labels, **trainparams)
    validation_generator = DataGenerator(partition['test'], labels, **testparams)

    # Model architecture
    model = tf.keras.models.Sequential()
    model.add(Input(shape=(128, 128, 70, 1)))
    # block 1
    model.add(Conv3D(8, kernel_size=3, activation='relu', padding='same'))
    model.add(MaxPool3D(pool_size=2))
    model.add(BatchNormalization())
    # block 2
    model.add(Conv3D(16, kernel_size=3, activation='relu', padding='same'))
    model.add(MaxPool3D(pool_size=2))
    model.add(BatchNormalization())
    # block 3
    model.add(Conv3D(32, kernel_size=3, activation='relu', padding='same'))
    model.add(MaxPool3D(pool_size=2))
    model.add(BatchNormalization())
    # block 4
    model.add(Conv3D(64, kernel_size=3, activation='relu', padding='same'))
    model.add(MaxPool3D(pool_size=2))
    model.add(BatchNormalization())
    # add additional block of Conv3D / MaxPool3D / BatchNorm with more filters for deeper network

    model.add(GlobalAveragePooling3D())
    model.add(Dense(128, activation='relu'))
    model.add(Dropout(0.5))
    model.add(Dense(128, activation='relu'))

    # Output Layer
    model.add(Dense(1, activation='sigmoid'))

    model.compile(optimizer=keras.optimizers.Adam(learning_rate=1e-3, decay=0.01),
                  loss="binary_crossentropy",
                  metrics=['accuracy'])

    # Optional early stopping
    # callback_stopping = EarlyStopping(monitor='val_accuracy', patience=10, restore_best_weights=True)

    # SAVE EPOCH WITH BEST VALIDATION ACCURACY
    callback_checkpoint = ModelCheckpoint(model_save_path + '/fold' + str(fold) + '_best',
                                          save_best_only=True,
                                          monitor='val_accuracy',
                                          mode='max')

    model.summary()

    # Train model on dataset
    history = model.fit(training_generator,
                        epochs=30,
                        validation_data=validation_generator,
                        shuffle=True,
                        callbacks=[callback_checkpoint])

    # Save final epoch 
    model.save(model_save_path + '/fold' + str(fold) + "_final")


    # Plot results
    fig, axs = plt.subplots(2)
    fig.suptitle('Network Loss and Accuracy')

    train_loss = history.history['loss']
    val_loss = history.history['val_loss']

    train_acc = history.history['accuracy']
    val_acc = history.history['val_accuracy']

    epochs = range(1, len(train_acc)+1)

    axs[0].plot(epochs, train_loss, 'bo', label='Training Loss')
    axs[0].plot(epochs, val_loss, 'b', label='Validation Loss')
    axs[0].set_ylabel('Loss [Binary Cross-Entropy]')

    axs[1].plot(epochs, train_acc, 'bo', label='Training Accuracy')
    axs[1].plot(epochs, val_acc, 'b', label='Validation Accuracy')
    axs[1].set_xlabel('Epochs')
    axs[1].set_ylabel('Accuracy')

    axs[0].legend(loc="upper right")
    plt.show()
